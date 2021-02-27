import os
from typing import (
    Optional,
    List,
    Tuple,
    Dict,
    Union,
    Any
)
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook


class ExcelWriter:
    """
    Excel sheet Writer class. Provide the writer method
    per column, row, cell

    :param path:
        location of xlsx file
    :param name:
        sheet name

    """

    def __init__(self, path, name=None):

        self._wb = Workbook(name)
        self._create_sheet()
        self.path = path
        self.row = 1
        self.col = 1

    @property
    def wb(self):
        return self._wb

    def write_col(self, values):
        """
        Method for writing value in sheet per column

        :param values(list):
            row values in order

        """
        for value in values:
            self.write_cell(self.row, self.col, value)
            self.row += 1
        self.row = 1
        self.col += 1

    def write_row(self, values: List[Any]):
        """
        Method for writing value in sheet per row

        :param values(list):
            row values in order

        """
        for value in values:
            self.write_cell(self.row, self.col, value)
            self.col += 1
        self.col = 1
        self.row += 1

    def write_cell(self, x, y, v):
        """
        Method for writing value in sheet per cell

        :param x:
            row location
        :param y:
            col location
        :param v:
            value to be entered

        """
        self.sheet.cell(x, y, v)

    def _create_sheet(self, name=None):
        """
        Method for creating new sheet. If name argument is not None,
        sheet name will be specified

        :param name:
            Sheet name

        """
        if name is None:
            sheet = self._wb.active
        else:
            sheet = self._wb.create_sheet(name)
        self.sheet = sheet

    def save(self):
        """
        Method for saving file in the directiory specified by developer.
        """
        self._wb.save(self.path)

class ExcelReader:
    """
    Excel reader class is iterator and read&parse file
    in row unit. Only read first sheet in file.

    """
    def __init__(self, path):
        self._wb = load_workbook(path)

    @property
    def active(self):
        """
        Return active sheet. the first sheet
        """
        return self._wb.active

    @property
    def header(self):
        """
        Return first row in sheet. Assume the first row is header
        and extract it
        """
        return self.rows[0]

    @property
    def body(self):
        """
        Everything except the first row.
        """
        return self.rows[1:]

    @property
    def rows(self):
        """
        Return all rows in sheet in the form of double list.

        :return(list[list]):
        """
        return list(self.active.rows)
    @property
    def size(self):
        """
        The number of rows except the first row
        """
        return len(self.body)

    def __iter__(self):
        self.idx = 0
        return self

    def __next__(self):
        if self.idx >= self.size:
            raise StopIteration
        r = [cell.value for cell in self.body[self.idx]]
        self.idx += 1
        return r

class _BaseAdater:
    """
    This class is Adapter base class which provide sink&exporter
    interface to executor. so If target document formation is changed
    and parsing logic is changed, just add custom adapter class
    that inherit this base class for parsing.

    :param path
        location of file
    :param output_model
        pydantic output model which deserialize response
        into desired response form
    """
    def __init__(self, path, output_model):
        self.path = path
        self._output_model = output_model

    def sink(self):
        """
        Method for reading file and extracting data
        from desired target directory for bulk data
        insert.
        """
        raise NotImplementedError

    def export(self, data):
        """
        Method for processing the data to create a file containing
        the desired data and shaped desired format. And drop the
        the file in desired location.

        :param data:
            data which will be processed

        """
        raise NotImplementedError


class SynonymExcelAdater(_BaseAdater):


    def __init__(self, path, output_model):
        super().__init__(path,  output_model)


    def sink(self):

        #Instance of Excel reader
        reader = ExcelReader(self.path)
        output_model = self._output_model
        synonym_dic = {}

        # _      : Serial number
        # k(str) : origin keyword
        # s(list): synonyms
        for _, k, s in reader:

            if k not in synonym_dic:
                synonym_dic[k] = []
            synonym_dic[k].append(s)

        response = [output_model(
                            origin_keyword=k,
                            synm_keyword=s).dict()
                    for k, s in synonym_dic.items()
                    ]
        return response

    def export(self, data):
        wirter = ExcelWriter(self.path)

        # First row is header
        wirter.write_row(
            ['number', 'keyword', 'synonym']
        )

        count = 1

        # data can not be generalized
        # because it depends on service that requires
        # desired data format
        for d in data:
            keyword = d['origin_keyword']
            synonyms = d['synonym']
            for synonym in synonyms:
                synonym = synonym['synm_keyword']
                wirter.write_row([count, keyword, synonym])

                # count mean the number header
                count += 1
        wirter.save()

class SynonymTextAdapter(_BaseAdater):

    def __init__(self, path, output_model):
        super().__init__(path,  output_model)

    def sink(self):
        output_model = self._output_model
        with open(self.path, 'r', encoding='utf-8') as file:

            # list of line
            lines = file.readlines()
            response = []

            # line is cut in unit of newline character
            # line: 'synonym1,synonym2,synonym3=>origin_keyword'
            for line in lines:

                # collapse line into synonyms and origin_keyword
                # synonyms(list)
                # keyword(str)
                synonyms, keyword = self._collapse_line(line)
                response.append(
                    output_model(
                            origin_keyword=keyword,
                            synm_keyword=synonyms).dict()
                )
        return response

    def export(self, datum):
        """

        :param datum(list):
            data which will be processed
            ex)
                datum = [{
                            "origin_keyword": k1,
                                    .
                                    .
                                    .

                            "synms":[
                                    {
                                        "synm_keyword": s1
                                    },
                                    .
                                    .
                                    .
                                    ]
                        }
                        .
                        .
                        .
                        ]
        """
        with open(self.path, 'w', encoding='utf-8') as file:
            for data in datum:

                # convert data in the form 'synonym1,synonym2,...=>origin_keyword'
                line = self._stick_line(data)
                file.write(line + '\n')

    def _stick_line(self, data) -> str:
        keyword = data['origin_keyword']
        synonyms = data['synms']
        synonyms = [synonym['synm_keyword'] for synonym in synonyms]
        sticked_line = ','.join(synonyms) + '=>' + keyword
        return sticked_line

    def _collapse_line(self, line: str) -> Tuple[List[str], str]:
        synonyms, keyword = line.split('=>')
        keyword = keyword.rstrip()
        synonyms = synonyms.split(',')
        return synonyms, keyword


class _BaseExecutor:
    """
    This class is adapter factory class and execute adapter
    according to its purpose(sink or exporter). And provide
    single interface 'def run()...' to file handler. So very
    easilly plugin a handler using def run()... interface.

    :param adapters(dict)
        Containing Identifier and adapters
    :param path
        target file path
    :param(Optional)
        data to export
    """

    adapters = {
        ".xlsx": SynonymExcelAdater,
        ".txt": SynonymTextAdapter
    }

    def __init__(self, path, data):
        self.path = path
        self.data = data

    def run(self):
        """
        Method for executing adapter
        :return:
        """
        raise NotImplementedError

class Exporter(_BaseExecutor):

    def __init__(self, path, ext, data=None, output_model=None):
        super().__init__(path, data)
        self.exporter = self.adapters[ext](path, output_model)

    def run(self):

        # run exporter
        response = self.exporter.export(self.data)
        return response

class Sinker(_BaseExecutor):

    def __init__(self,
                 path,
                 ext,
                 data=None,
                 output_model=None):

        super().__init__(path, data)
        self.sinker = self.adapters[ext](path, output_model)

    def run(self):

        # run sinker
        # from file to db for bulk inserting
        response = self.sinker.sink()
        return response

class FileParser:

    """
    This is integrated file handler. It is very simple.
    Fist, specify path, executor class(sinker, exporter),
    data and output_model. Then, executor class as explanning
    above make a adapter according to purpose.
    so we just have to objectify handler and call process method.

    ===== In case of sink
    parser = FileParser(path=..., executor_class=Sinker, output_model=...)
    parser.process()

    ==== In case of exporter
    parser = FileParser(path=..., executor_class=Exporter, data=...., output_model=...)
    parser.process()

    :param path
        target file path.
        It has different meanning depending on whether is is exporter or sinker
        In the case of a sinker, it means the location of the directory where
        the file is stored but exporter, it means the location of the directory to save
    :param executor_class
        Sinker or Exporter
    :param data
        It is only specified when Export the data in file
    :param output_model
        pydantic model

    """
    def __init__(self,
                 path: str,
                 executor_class: Optional[Union[Sinker, Exporter]],
                 data: Optional[List[Dict[str, Any]]]=None,
                 output_model: Optional[BaseModel]=None
                 ):

        _, ext = os.path.splitext(path)
        if os.path.isdir(path) or ext not in ('.txt', '.xlsx'):
            raise ValueError('%s value error' % path,
                             'File must be <txt or excel> and <not directory>')

        self._path = path
        self.handler = executor_class(path, ext, data, output_model)

    @property
    def path(self):
        return self._path

    def process(self):
        """
        Process the exporter or Sinker in single interface.
        """

        response = self.handler.run()
        return response

    def remove(self):
        """
        Remove the file when file is not needed after processing
        """

        os.remove(self.path)